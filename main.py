import PySimpleGUI as sg
from PIL import Image
from PIL import ImageOps
import openpyxl as px
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def rgb_to_hex(r, g, b):
    return 'FF%02x%02x%02x' % (r, g, b)


if __name__ == '__main__':

    selected_file = None
    open = True

    sg.theme("DarkTeal2")
    layout = [[], [sg.Text("Choose a file: "), sg.Input(), sg.FileBrowse(key="-IN-")], [sg.Button("Submit")]]

    window = sg.Window('My File Browser', layout, size=(600, 90))

    while open:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == "Exit":
            open = False
        elif event == "Submit":
            selected_file = values["-IN-"]
            open = False

    try:
        MAX_SIDE = 250
        MIN_SIDE = 125
        with Image.open(selected_file) as img:

            img = ImageOps.exif_transpose(img)
            width, height = img.size

            # print(width, height)

            w_ratio = h_ratio = 1
            if width > MAX_SIDE or height > MAX_SIDE:
                if width > height:
                    height = int((height / width * MAX_SIDE) // 1)
                    width = MAX_SIDE
                else:
                    width = int((width / height * MAX_SIDE) // 1)
                    height = MAX_SIDE

            # print(width, height)

            if width < MIN_SIDE or height < MIN_SIDE:
                if width < height:
                    w_ratio = MIN_SIDE / width
                    width = MIN_SIDE
                else:
                    h_ratio = MIN_SIDE / height
                    height = MIN_SIDE

            # print(width, height)

            img = img.resize((width, height))

            rgb_im = img.convert('RGB')

            # rgb_im.show()

            # print(w_ratio, h_ratio)

            wb = px.Workbook()
            ws = wb.worksheets[0]
            BASE_SIZE = 0.5
            R_TO_C = 5.69   # adjustment for differences in row / col sizing in Excel
            for c in range(1, width+1):
                coldim = ws.column_dimensions[get_column_letter(c)]
                coldim.width = BASE_SIZE * w_ratio
            for r in range(1, height+1):
                rowdim = ws.row_dimensions[r]
                rowdim.height = BASE_SIZE * h_ratio * R_TO_C

            def fill_cell(cell, color):
                cell.fill = PatternFill("solid", start_color=color)
                pass

            for c in range(width):
                for r in range(height):
                    fill_cell(
                        ws.cell(row=r+1, column=c+1),
                        rgb_to_hex(*rgb_im.getpixel((c, r)))
                    )

            ws.sheet_view.zoomScale=75
            wb.save("file.xlsx")

    except Exception as ex:
        print(f"Error:\n{ex}")